from flask import Blueprint, render_template, request, jsonify, current_app
from flask_login import login_required, current_user
from app.models import db, User, CertificateType, AuditLog, CertArchive, EmailDraft, OrgSettings, Campaign
from app.engine.cert_id import assign_certificate_id, verify_format
from app.engine.ocr_analyzer import analyze_template
from datetime import datetime
import uuid
import os
import json
import io
import csv
import hashlib
from app.models import CertificateAsset
bp = Blueprint('certificates', __name__)


@bp.route('/admin')
@login_required
def dashboard():
    cert_types = CertificateType.query.filter_by(is_active=True).order_by(CertificateType.created_at.desc()).all()
    org = OrgSettings.query.first() or OrgSettings()
    drafts = EmailDraft.query.order_by(EmailDraft.updated_at.desc()).all()
    return render_template('admin/dashboard.html', cert_types=cert_types, org=org, drafts=drafts)


@bp.route('/admin/truth/certificates/<cert_id>')
@login_required
def resolve_certificate_truth(cert_id):
    from app.engine.clp import resolve_truth
    result = resolve_truth(cert_id)
    return jsonify(result)


@bp.route('/api/org', methods=['GET'])
@login_required
def get_org():
    org = OrgSettings.query.first() or OrgSettings()
    return jsonify({'org_name': org.org_name, 'sender_name': org.sender_name, 'sender_email': org.sender_email, 'reply_to_email': org.reply_to_email, 'verify_base_url': org.verify_base_url})


@bp.route('/api/org', methods=['POST'])
@login_required
def save_org():
    data = request.json or {}
    org = OrgSettings.query.first()
    if not org:
        org = OrgSettings()
        db.session.add(org)
    for field in ('org_name', 'sender_name', 'sender_email', 'reply_to_email', 'verify_base_url'):
        if field in data:
            setattr(org, field, data[field])
    db.session.commit()
    return jsonify({'message': 'Settings saved'})


@bp.route('/api/cert-types', methods=['GET'])
@login_required
def list_cert_types():
    types = CertificateType.query.filter_by(is_active=True).order_by(CertificateType.created_at.desc()).all()
    return jsonify([{'id': ct.id, 'name': ct.name, 'course_code': ct.course_code, 'period': ct.period, 'registration_token': ct.registration_token, 'master_file_type': ct.master_file_type, 'user_count': len(ct.users), 'seq_counter': ct.seq_counter} for ct in types])


@bp.route('/api/cert-types', methods=['POST'])
@login_required
def create_cert_type():
    file = request.files.get('master_file')
    if not file:
        return (jsonify({'error': 'Master file required'}), 400)
    name = request.form.get('name', '').strip()
    period = request.form.get('period', '').strip()
    course_code = request.form.get('course_code', 'GEN').strip().upper()
    if not name or not period:
        return (jsonify({'error': 'Name and period required'}), 400)
    ext = file.filename.rsplit('.', 1)[-1].lower()
    if ext not in ('pdf', 'png'):
        return (jsonify({'error': 'Only PDF or PNG allowed'}), 400)
    
    master_binary = file.read()
    file_hash = hashlib.md5(master_binary).hexdigest()
    asset_id = str(uuid.uuid4())
    
    asset = CertificateAsset(
        id=asset_id,
        name=file.filename,
        file_binary=master_binary,
        file_hash=file_hash
    )
    db.session.add(asset)
    db.session.flush()

    overlay_coords = {'name_x': 150, 'name_y': 380, 'name_w': 500, 'name_h': 50, 'name_font_size': 28, 'name_align': 'center', 'cert_id_x': 400,
                      'cert_id_y': 80, 'cert_id_font_size': 9, 'date_x': 150, 'date_y': 200, 'date_font_size': 9, 'qr_x': 0, 'qr_y': 30, 'qr_size': 90}
    try:
        custom = json.loads(request.form.get('overlay_coords') or 'null')
        if custom:
            overlay_coords.update(custom)
    except Exception:
        pass
        
    ocr_result = {'regions': [], 'message': 'OCR skipped for DB asset'}
    token = str(uuid.uuid4())[:8]
    
    ct = CertificateType(name=name, course_code=course_code, period=period, asset_id=asset_id, master_pdf_path="", master_pdf_binary=None, master_file_type=ext, overlay_coords=overlay_coords, ocr_regions=ocr_result.get(
        'regions'), registration_token=token, email_message=request.form.get('email_message', '').strip() or None, email_subject=request.form.get('email_subject', '').strip() or None, seq_counter=0)
    db.session.add(ct)
    db.session.commit()

    # Automatically convert newly uploaded PDF template to SVG if possible
    if ext == 'pdf':
        try:
            os.makedirs("uploads", exist_ok=True)
            pdf_path = f"uploads/{ct.id}_master.pdf"
            with open(pdf_path, 'wb') as f:
                f.write(master_binary)
            
            from app.engine.template_converter import pdf_to_svg, add_placeholders_to_svg
            svg_path = f"uploads/{ct.id}_master.svg"
            pdf_to_svg(pdf_path, svg_path)
            add_placeholders_to_svg(svg_path, ct.overlay_coords)
            
            ct.master_svg_path = svg_path
            db.session.commit()
            if os.path.exists(pdf_path):
                os.remove(pdf_path)
        except Exception as e:
            # Safe silent fallback for robust operation
            pass

    return jsonify({'id': ct.id, 'name': ct.name, 'registration_token': token, 'register_url': f'/register/{token}', 'ocr_message': ocr_result.get('message'), 'ocr_regions': ocr_result.get('regions')})


@bp.route('/api/cert-types/<int:ct_id>', methods=['GET'])
@login_required
def get_cert_type(ct_id):
    ct = CertificateType.query.get_or_404(ct_id)
    return jsonify({'id': ct.id, 'name': ct.name, 'course_code': ct.course_code, 'period': ct.period, 'master_file_type': ct.master_file_type, 'overlay_coords': ct.overlay_coords, 'ocr_regions': ct.ocr_regions, 'email_subject': ct.email_subject, 'email_message': ct.email_message, 'registration_token': ct.registration_token})


@bp.route('/api/cert-types/<int:ct_id>', methods=['PUT'])
@login_required
def update_cert_type(ct_id):
    ct = CertificateType.query.get_or_404(ct_id)
    data = request.json or {}
    for field in ('name', 'period', 'course_code', 'overlay_coords', 'email_message', 'email_subject', 'ocr_regions'):
        if field in data:
            setattr(ct, field, data[field])
    db.session.commit()
    return jsonify({'message': 'Updated'})


@bp.route('/api/cert-types/<int:ct_id>', methods=['DELETE'])
@login_required
def delete_cert_type(ct_id):
    ct = CertificateType.query.get_or_404(ct_id)
    ct.is_active = False
    db.session.commit()
    return jsonify({'message': 'Removed from library'})


@bp.route('/admin/upload-template', methods=['POST'])
@login_required
def upload_template():
    """Admin uploads PDF master → system converts to SVG template"""
    if 'master_pdf' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    file = request.files['master_pdf']
    cert_type_id = request.form.get('cert_type_id')
    if not cert_type_id:
        return jsonify({'error': 'cert_type_id required'}), 400
    
    cert_type = CertificateType.query.get(cert_type_id)
    if not cert_type:
        return jsonify({'error': 'CertificateType not found'}), 404
    
    # Ensure upload folder exists
    os.makedirs("uploads", exist_ok=True)
    
    # Save uploaded PDF temporarily
    pdf_path = f"uploads/{cert_type_id}_master.pdf"
    file.save(pdf_path)
    
    # Convert to SVG
    from app.engine.template_converter import pdf_to_svg, add_placeholders_to_svg
    
    svg_path = f"uploads/{cert_type_id}_master.svg"
    try:
        pdf_to_svg(pdf_path, svg_path)
        
        # Add placeholders based on overlay_coords
        add_placeholders_to_svg(svg_path, cert_type.overlay_coords)
        
        # Update certificate type to use SVG
        cert_type.master_svg_path = svg_path
        db.session.commit()
    finally:
        # Clean up PDF if it was saved
        if os.path.exists(pdf_path):
            os.remove(pdf_path)
    
    return jsonify({'message': 'Template converted to SVG', 'svg_path': svg_path})


@bp.route('/api/cert-types/<int:ct_id>/analyze', methods=['POST'])
@login_required
def reanalyze_template(ct_id):
    ct = CertificateType.query.get_or_404(ct_id)
    return jsonify({'regions': [], 'message': 'OCR disabled for DB assets'})


@bp.route('/api/cert-types/<int:ct_id>/preview', methods=['POST'])
@login_required
def preview_certificate(ct_id):
    from flask import send_file
    from app.engine.pdf_processor import generate_personalized_pdf
    from app.domain.certificates.service import resolve_certificate_asset, ensure_svg_template
    ct = CertificateType.query.get_or_404(ct_id)
    data = request.json or {}
    full_name = data.get('full_name', 'Jane Smith')
    cert_id = data.get('cert_id', 'MLJ-GEN-2026-000001')
    issue_date = data.get('issue_date', datetime.utcnow().strftime('%d %B %Y'))
    include_qr = data.get('include_qr', True)
    try:
        if ct.master_svg_path:
            template_source = ensure_svg_template(ct) or resolve_certificate_asset(ct)
        else:
            template_source = resolve_certificate_asset(ct)
        pdf_bytes = generate_personalized_pdf(template_source, overlay_coords=ct.overlay_coords, full_name=full_name,
                                              certificate_id=cert_id, issuance_date=issue_date, include_qr=include_qr, cert_name=ct.name, master_file_type=ct.master_file_type or 'pdf')
    except RuntimeError as e:
        return (jsonify({'error': str(e)}), 404)
    except Exception as e:
        return (jsonify({'error': f'Failed to generate preview: {str(e)}'}), 500)
    return send_file(io.BytesIO(pdf_bytes), mimetype='application/pdf', as_attachment=False, download_name=f'preview_{ct_id}.pdf')


@bp.route('/api/users/<int:cert_type_id>')
@login_required
def get_users(cert_type_id):
    status_filter = request.args.get('status')
    q = User.query.filter_by(certificate_type_id=cert_type_id)
    if status_filter:
        q = q.filter_by(status=status_filter)
    users = q.order_by(User.created_at.desc()).all()
    return jsonify({'data': [{'id': u.id, 'full_name': u.full_name, 'email': u.email, 'certificate_id': u.certificate_id or '—', 'status': u.status, 'include_qr': u.include_qr, 'score': u.score, 'source': u.source, 'created_at': u.created_at.strftime('%d/%m/%Y %H:%M') if u.created_at else ''} for u in users]})


@bp.route('/api/users/add', methods=['POST'])
@login_required
def add_user():
    data = request.json or {}
    ct_id = data.get('cert_type_id')
    if not ct_id:
        return (jsonify({'error': 'cert_type_id required'}), 400)
    ct = CertificateType.query.get_or_404(ct_id)
    first_name = (data.get('first_name') or '').strip()
    surname = (data.get('surname') or '').strip()
    email = (data.get('email') or '').strip().lower()
    if not first_name or not surname or (not email):
        return (jsonify({'error': 'first_name, surname and email required'}), 400)
    user = User(first_name=first_name, surname=surname, other_name=(data.get('other_name') or '').strip(),
                email=email, certificate_type_id=ct_id, status='registered', source='manual')
    db.session.add(user)
    db.session.flush()
    cert_id = assign_certificate_id(user)
    db.session.add(AuditLog(action='registered', performed_by=current_user.email,
                   details={'cert_id': cert_id, 'email': email}))
    db.session.commit()
    return jsonify({'message': 'Participant added', 'id': user.id, 'certificate_id': cert_id})


@bp.route('/api/users/import', methods=['POST'])
@login_required
def import_users():
    import openpyxl
    ct_id = request.form.get('cert_type_id')
    ct = CertificateType.query.get_or_404(ct_id)
    file = request.files.get('file')
    if not file:
        return (jsonify({'error': 'File required'}), 400)
    ext = file.filename.rsplit('.', 1)[-1].lower()
    rows = []
    if ext == 'csv':
        content = file.read().decode('utf-8', errors='ignore')
        reader = csv.DictReader(io.StringIO(content))
        rows = list(reader)
    elif ext in ('xlsx', 'xls'):
        wb = openpyxl.load_workbook(io.BytesIO(file.read()))
        ws = wb.active
        headers = [str(c.value or '').strip().lower() for c in next(ws.iter_rows(min_row=1, max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append(dict(zip(headers, row)))
    else:
        return (jsonify({'error': 'CSV or XLSX only'}), 400)
    imported = skipped = 0
    for row in rows:
        email = str(row.get('email') or row.get('Email') or '').strip().lower()
        first_name = str(row.get('first_name') or row.get('First Name') or row.get('firstname') or '').strip()
        surname = str(row.get('surname') or row.get('last_name') or row.get(
            'Last Name') or row.get('lastname') or '').strip()
        if not email or not first_name:
            skipped += 1
            continue
        if not first_name and (not surname):
            full = str(row.get('full_name') or row.get('name') or '')
            parts = full.split()
            first_name = parts[0] if parts else ''
            surname = parts[-1] if len(parts) > 1 else ''
        user = User(first_name=first_name, surname=surname, email=email, certificate_type_id=int(ct_id), status='registered', score=float(row.get(
            'score') or 0) if row.get('score') else None, completion_status=str(row.get('completion_status') or '').strip() or None, source='csv')
        db.session.add(user)
        db.session.flush()
        assign_certificate_id(user)
        imported += 1
    db.session.commit()
    return jsonify({'message': f'{imported} imported, {skipped} skipped'})


@bp.route('/api/users/toggle-qr', methods=['POST'])
@login_required
def toggle_qr():
    uid = request.json.get('user_id')
    user = db.session.get(User, uid)
    if not user:
        return (jsonify({'error': 'Not found'}), 404)
    user.include_qr = not user.include_qr
    db.session.commit()
    return jsonify({'include_qr': user.include_qr})


@bp.route('/api/approve', methods=['POST'])
@login_required
def approve_users():
    """Approve users AND generate certificates immediately. PDF saved to archive."""
    data = request.json or {}
    user_ids = data.get('user_ids', [])
    cert_type_id = data.get('cert_type_id')

    if not user_ids:
        return jsonify({'error': 'No users selected'}), 400

    users = User.query.filter(
        User.id.in_(user_ids),
        User.status.in_(['registered', 'rejected'])
    ).all()

    if not users:
        return jsonify({'error': 'No eligible users found in selection'}), 400

    # Determine cert_type: from payload, or from the first user's cert type
    cert_type = None
    if cert_type_id:
        cert_type = CertificateType.query.get(cert_type_id)
    if not cert_type and users:
        cert_type = users[0].certificate_type
    if not cert_type:
        return jsonify({'error': 'Certificate type not found'}), 404

    from app.engine.pdf_processor import generate_personalized_pdf
    from app.domain.certificates.service import resolve_certificate_asset, ensure_svg_template
    from app.engine.cert_id import assign_certificate_id
    from app.models import Certificate, CertificateStatus, OrgSettings

    org = OrgSettings.query.first() or OrgSettings()
    base_url = (org.verify_base_url or '').rstrip('/')

    # Resolve the template source once for all users
    if cert_type.master_svg_path:
        template_source = ensure_svg_template(cert_type) or resolve_certificate_asset(cert_type)
    else:
        template_source = resolve_certificate_asset(cert_type)

    generated = 0
    failed = 0
    preview_urls = []

    for user in users:
        try:
            # 1. Assign certificate ID if not yet assigned
            if not user.certificate_id:
                assign_certificate_id(user)
                db.session.flush()

            full_name = f"{user.first_name} {user.surname}"
            if user.other_name and user.other_name.strip():
                full_name = f"{full_name} {user.other_name}"
            full_name = full_name.strip()

            issue_date = datetime.utcnow().strftime('%d %B %Y')
            verify_url = f"{base_url}/verify/{user.certificate_id}" if base_url else ''

            # 2. Generate the PDF
            pdf_bytes = generate_personalized_pdf(
                template_source,
                overlay_coords=cert_type.overlay_coords,
                full_name=full_name,
                certificate_id=user.certificate_id,
                issuance_date=issue_date,
                include_qr=user.include_qr,
                cert_name=cert_type.name,
                master_file_type=cert_type.master_file_type or 'pdf',
                verify_url=verify_url
            )

            # 3. Save to archive directory
            os.makedirs('archive', exist_ok=True)
            archive_path = f"archive/{user.certificate_id}.pdf"
            with open(archive_path, 'wb') as f:
                f.write(pdf_bytes)

            # 4. Ensure Certificate record is up to date
            import hashlib
            pdf_hash = hashlib.sha256(pdf_bytes).hexdigest()
            cert = Certificate.query.get(user.certificate_id)
            if not cert:
                cert = Certificate(
                    id=user.certificate_id,
                    user_id=user.id,
                    cert_type_id=cert_type.id,
                    asset_id=cert_type.asset_id,
                    status=CertificateStatus.DRAFT
                )
                db.session.add(cert)
                db.session.flush()

            # Transition through states to READY_FOR_DISPATCH
            if cert.status == CertificateStatus.DRAFT:
                cert.transition_to_approved()
            if cert.status == CertificateStatus.APPROVED_FOR_GENERATION:
                cert.transition_to_generating()
            cert.transition_to_generated(pdf_bytes, pdf_hash)
            cert.transition_to_ready()

            # 5. Approve the user
            user.status = 'approved'
            user.approved_at = datetime.utcnow()

            db.session.add(AuditLog(
                user_id=user.id,
                action='approved',
                performed_by=current_user.email,
                details={
                    'certificate_generated': True,
                    'archive_path': archive_path,
                    'file_size_bytes': len(pdf_bytes)
                }
            ))

            preview_urls.append(f"/api/archive/view/{user.certificate_id}")
            generated += 1

        except Exception as e:
            current_app.logger.error(f"Failed to approve/generate for user {user.id}: {e}", exc_info=True)
            db.session.add(AuditLog(
                user_id=user.id,
                action='approve_failed',
                performed_by=current_user.email,
                details={'error': str(e)}
            ))
            # Still mark user approved even if PDF generation failed
            user.status = 'approved'
            user.approved_at = datetime.utcnow()
            failed += 1

    db.session.commit()

    msg = f'{generated} approved with certificates generated'
    if failed:
        msg += f', {failed} approved (PDF generation failed — will retry on Send)'

    return jsonify({
        'message': msg,
        'generated': generated,
        'failed': failed,
        'preview_urls': preview_urls[:5]
    })


@bp.route('/api/reject', methods=['POST'])
@login_required
def reject_users():
    user_ids = request.json['user_ids']
    User.query.filter(User.id.in_(user_ids)).update({'status': 'rejected'}, synchronize_session=False)
    db.session.commit()
    return jsonify({'message': f'{len(user_ids)} rejected'})


@bp.route('/api/send', methods=['POST'])
@login_required
def send_certificates():
    data = request.json or {}
    user_ids = data['user_ids']
    cert_type_id = data['cert_type_id']
    draft_id = data.get('draft_id')
    confirmed = data.get('confirmed', False)
    if not confirmed:
        users = User.query.filter(User.id.in_(user_ids), User.status == 'approved').all()
        ct = CertificateType.query.get(cert_type_id)
        org = OrgSettings.query.first() or OrgSettings()
        return jsonify({'needs_confirmation': True, 'recipient_count': len(users), 'course': ct.name if ct else '', 'sender': f"{org.sender_name} <{org.sender_email or current_app.config.get('MAIL_USERNAME', '')}>", 'cert_type_id': cert_type_id})
    ct = CertificateType.query.get(cert_type_id)
    if draft_id:
        draft = db.session.get(EmailDraft, draft_id)
        if not draft:
            return (jsonify({'error': 'Selected email draft not found.'}), 400)
    elif not ct.email_message:
        default_draft = EmailDraft.query.filter_by(is_default=True).first()
        if not default_draft:
            return (jsonify({'error': 'No email template configured. Please provide a message in the Certificate Type settings or create a default Email Draft.'}), 400)
    from app.models import Certificate, CertificateStatus
    users = User.query.filter(User.id.in_(user_ids), User.status == 'approved').all()
    queued = 0
    for user in users:
        # Ensure Certificate record exists (IDEMPOTENT)
        if not user.certificate_id:
            assign_certificate_id(user)
            db.session.commit()

        cert = Certificate.query.get(user.certificate_id)
        if not cert:
            cert = Certificate(
                id=user.certificate_id,
                user_id=user.id,
                cert_type_id=cert_type_id,
                status=CertificateStatus.DRAFT
            )
            db.session.add(cert)
            cert._log_event("CERT_CREATED")
            db.session.commit()

        # Transition to APPROVED_FOR_GENERATION
        if cert.status == CertificateStatus.DRAFT:
            cert.transition_to_approved()
            db.session.commit()

        # Trigger generation flow
        if cert.status == CertificateStatus.APPROVED_FOR_GENERATION:
            current_app.task_queue.enqueue(
                'generate_certificate', 
                idempotency_key=f"cert_{user.id}_{cert_type_id}", 
                certificate_id=cert.id,
                draft_id=draft_id
            )
            queued += 1

    db.session.add(AuditLog(action='batch_send_initiated', performed_by=current_user.email,
                   details={'count': queued, 'cert_type_id': cert_type_id}))
    db.session.commit()
    return jsonify({'message': f'{queued} certificates queued for dispatch'})


@bp.route('/api/send-all-approved', methods=['POST'])
@login_required
def send_all_approved():
    data = request.json or {}
    cert_type_id = data['cert_type_id']
    draft_id = data.get('draft_id')
    confirmed = data.get('confirmed', False)
    users = User.query.filter_by(certificate_type_id=cert_type_id, status='approved').all()
    if not confirmed:
        ct = CertificateType.query.get(cert_type_id)
        org = OrgSettings.query.first() or OrgSettings()
        return jsonify({'needs_confirmation': True, 'recipient_count': len(users), 'course': ct.name if ct else '', 'sender': f"{org.sender_name} <{org.sender_email or current_app.config.get('MAIL_USERNAME', '')}>"})
    if not users:
        return jsonify({'message': 'No approved participants found'})
    ct = CertificateType.query.get(cert_type_id)
    if draft_id:
        draft = db.session.get(EmailDraft, draft_id)
        if not draft:
            return (jsonify({'error': 'Selected email draft not found.'}), 400)
    elif not ct.email_message:
        default_draft = EmailDraft.query.filter_by(is_default=True).first()
        if not default_draft:
            return (jsonify({'error': 'No email template configured. Please provide a message in the Certificate Type settings or create a default Email Draft.'}), 400)
    from app.models import Certificate, CertificateStatus
    queued = 0
    for user in users:
        # Ensure Certificate record exists
        if not user.certificate_id:
            assign_certificate_id(user)
            db.session.commit()

        cert = Certificate.query.get(user.certificate_id)
        if not cert:
            cert = Certificate(
                id=user.certificate_id,
                user_id=user.id,
                cert_type_id=cert_type_id,
                status=CertificateStatus.DRAFT
            )
            db.session.add(cert)
            cert._log_event("CERT_CREATED")
            db.session.commit()

        if cert.status == CertificateStatus.DRAFT:
            cert.transition_to_approved()
            db.session.commit()

        if cert.status == CertificateStatus.APPROVED_FOR_GENERATION:
            current_app.task_queue.enqueue(
                'generate_certificate', 
                idempotency_key=f"cert_{user.id}_{cert_type_id}", 
                certificate_id=cert.id,
                draft_id=draft_id
            )
            queued += 1

    db.session.add(AuditLog(action='send_all_approved', performed_by=current_user.email,
                   details={'count': queued, 'cert_type_id': cert_type_id}))
    db.session.commit()
    return jsonify({'message': f'{queued} certificates queued for dispatch'})


@bp.route('/api/nudge', methods=['POST'])
@login_required
def nudge_user():
    uid = request.json['user_id']
    current_app.task_queue.enqueue('send_nudge_email', idempotency_key=f"nudge_{uid}_{datetime.utcnow().strftime('%Y%m%d')}", user_id=uid, job_timeout=30)
    return jsonify({'message': 'Reminder queued'})


@bp.route('/api/delete', methods=['POST'])
@login_required
def delete_users():
    user_ids = request.json['user_ids']
    users = User.query.filter(User.id.in_(user_ids)).all()
    for u in users:
        if u.certificate_id and (not CertArchive.query.filter_by(certificate_id=u.certificate_id).first()):
            db.session.add(CertArchive(certificate_id=u.certificate_id, full_name=u.full_name, cert_name=u.certificate_type.name if u.certificate_type else '', course_code=u.certificate_type.course_code if u.certificate_type else '',
                           issued_date=u.sent_at.strftime('%d %B %Y') if u.sent_at else '', email=u.email, status='archived', raw_binary=json.dumps({'email': u.email}).encode('utf-8')))
    User.query.filter(User.id.in_(user_ids)).delete(synchronize_session=False)
    db.session.add(AuditLog(action='deleted', performed_by=current_user.email, details={'count': len(user_ids)}))
    db.session.commit()
    return jsonify({'message': f'{len(user_ids)} participants deleted'})


@bp.route('/api/users/archive', methods=['POST'])
@login_required
def archive_users():
    user_ids = request.json['user_ids']
    User.query.filter(User.id.in_(user_ids)).update(
        {'archived_at': datetime.utcnow(), 'status': 'archived'}, synchronize_session=False)
    db.session.commit()
    return jsonify({'message': f'{len(user_ids)} participants archived'})


@bp.route('/api/drafts', methods=['GET'])
@login_required
def list_drafts():
    drafts = EmailDraft.query.order_by(EmailDraft.updated_at.desc()).all()
    return jsonify([{'id': d.id, 'name': d.name, 'subject': d.subject, 'body': d.body, 'include_attachment': d.include_attachment, 'is_default': d.is_default, 'updated_at': d.updated_at.strftime('%d/%m/%Y') if d.updated_at else ''} for d in drafts])


@bp.route('/api/drafts', methods=['POST'])
@login_required
def save_draft():
    data = request.json or {}
    draft = EmailDraft(name=data.get('name', 'Untitled Draft'), subject=data.get('subject', 'Your {{course_name}} Certificate'), body=data.get(
        'body', ''), include_attachment=data.get('include_attachment', True), is_default=data.get('is_default', False))
    db.session.add(draft)
    db.session.commit()
    return jsonify({'id': draft.id, 'message': 'Draft saved'})


@bp.route('/api/drafts/<int:did>', methods=['PUT'])
@login_required
def update_draft(did):
    draft = db.session.get(EmailDraft, did)
    if not draft:
        return (jsonify({'error': 'Not found'}), 404)
    data = request.json or {}
    for f in ('name', 'subject', 'body', 'include_attachment', 'is_default'):
        if f in data:
            setattr(draft, f, data[f])
    draft.updated_at = datetime.utcnow()
    db.session.commit()
    return jsonify({'message': 'Draft updated'})


@bp.route('/api/drafts/<int:did>', methods=['DELETE'])
@login_required
def delete_draft(did):
    draft = db.session.get(EmailDraft, did)
    if not draft:
        return (jsonify({'error': 'Not found'}), 404)
    db.session.delete(draft)
    db.session.commit()
    return jsonify({'message': 'Draft deleted'})


@bp.route('/api/campaigns', methods=['GET'])
@login_required
def list_campaigns():
    camps = Campaign.query.order_by(Campaign.created_at.desc()).all()
    return jsonify([{'id': c.id, 'name': c.name, 'status': c.status, 'recipient_count': c.recipient_count, 'sent_count': c.sent_count, 'failed_count': c.failed_count, 'include_attachment': c.include_attachment, 'created_at': c.created_at.strftime('%d/%m/%Y') if c.created_at else '', 'sent_at': c.sent_at.strftime('%d/%m/%Y %H:%M') if c.sent_at else ''} for c in camps])


@bp.route('/api/campaigns', methods=['POST'])
@login_required
def create_campaign():
    data = request.json or {}
    user_ids = data.get('user_ids', [])
    draft_id = data.get('draft_id')
    cert_type_id = data.get('cert_type_id')
    send_now = data.get('send_now', False)
    campaign = Campaign(name=data.get('name', f"Campaign {datetime.utcnow().strftime('%d/%m/%Y')}"), cert_type_id=cert_type_id, draft_id=draft_id,
                        include_attachment=data.get('include_attachment', False), recipient_count=len(user_ids), status='draft', created_by=current_user.email)
    db.session.add(campaign)
    db.session.commit()
    if send_now:
        current_app.task_queue.enqueue('process_campaign', idempotency_key=f"campaign_{campaign.id}", campaign_id=campaign.id,
                                       user_ids=user_ids, draft_id=draft_id, job_timeout=600)
        campaign.status = 'scheduled'
        db.session.commit()
    return jsonify({'id': campaign.id, 'message': 'Campaign created'})


@bp.route('/api/lms/completion', methods=['POST'])
def lms_completion():
    data = request.json or {}
    email = (data.get('email') or data.get('learner_email') or '').strip().lower()
    full_name = (data.get('full_name') or data.get('name') or '').strip()
    score = data.get('score') or data.get('grade')
    completion = data.get('completion_status') or data.get('status') or 'completed'
    token = data.get('registration_token') or request.args.get('token', '')
    if not email or not token:
        return (jsonify({'error': 'email and registration_token required'}), 400)
    ct = CertificateType.query.filter_by(registration_token=token, is_active=True).first()
    if not ct:
        return (jsonify({'error': 'Invalid token'}), 404)
    name_parts = full_name.split()
    if len(name_parts) == 1:
        first_name = name_parts[0]
        surname = ''
    elif len(name_parts) > 1:
        first_name = name_parts[0]
        surname = ' '.join(name_parts[1:])
    else:
        first_name = email.split('@')[0]
        surname = ''
    user = User(first_name=first_name, surname=surname, email=email, certificate_type_id=ct.id,
                status='registered', score=float(score) if score else None, completion_status=completion, source='lms')
    db.session.add(user)
    db.session.flush()
    cert_id = assign_certificate_id(user)
    db.session.commit()
    return jsonify({'message': 'Participant registered', 'certificate_id': cert_id, 'status': 'pending_admin_approval'})


@bp.route('/api/lms/users', methods=['POST'])
def lms_users():
    data = request.json or {}
    token = data.get('registration_token') or request.args.get('token', '')
    users_data = data.get('users', [])
    ct = CertificateType.query.filter_by(registration_token=token, is_active=True).first()
    if not ct:
        return (jsonify({'error': 'Invalid token'}), 404)
    imported = 0
    for u in users_data:
        email = (u.get('email') or '').strip().lower()
        if not email:
            continue
        full = (u.get('full_name') or u.get('name') or '').strip()
        name_parts = full.split()
        if len(name_parts) == 1:
            f_name = name_parts[0]
            s_name = ''
        elif len(name_parts) > 1:
            f_name = name_parts[0]
            s_name = ' '.join(name_parts[1:])
        else:
            f_name = email.split('@')[0]
            s_name = ''
        user = User(first_name=f_name, surname=s_name, email=email,
                    certificate_type_id=ct.id, status='registered', source='lms')
        db.session.add(user)
        db.session.flush()
        assign_certificate_id(user)
        imported += 1
    db.session.commit()
    return jsonify({'message': f'{imported} users synced'})


@bp.route('/verify/<cert_id>')
def verify_cert(cert_id):
    record = CertArchive.query.filter_by(certificate_id=cert_id).first()
    if not record:
        user = User.query.filter_by(certificate_id=cert_id).first()
        if user:
            record = type('R', (), {'certificate_id': user.certificate_id, 'full_name': user.full_name, 'cert_name': user.certificate_type.name if user.certificate_type else '',
                          'issued_date': user.sent_at.strftime('%d %B %Y') if user.sent_at else '', 'status': user.status})()
    fmt_valid = verify_format(cert_id)
    return render_template('public/verify.html', record=record, fmt_valid=fmt_valid, cert_id=cert_id)


@bp.route('/api/archive/<cert_id>')
@login_required
def get_archive(cert_id):
    record = CertArchive.query.filter_by(certificate_id=cert_id).first()
    if not record:
        return (jsonify({'error': 'Not found'}), 404)
    data = record.to_dict()
    if record.raw_binary:
        try:
            data['raw'] = json.loads(record.raw_binary.decode('utf-8'))
        except Exception:
            pass
    return jsonify(data)


@bp.route('/api/audit')
@login_required
def get_audit():
    logs = AuditLog.query.order_by(AuditLog.created_at.desc()).limit(200).all()
    return jsonify([{'action': l.action, 'performed_by': l.performed_by, 'details': l.details, 'created_at': l.created_at.strftime('%d/%m/%Y %H:%M:%S') if l.created_at else ''} for l in logs])


@bp.route('/api/archive/view/<cert_id>')
def view_certificate(cert_id):
    """
    Public/Admin certificate preview endpoint.
    Shows the actual generated PDF in browser.
    """
    from flask import send_file, jsonify
    user = User.query.filter_by(certificate_id=cert_id).first()
    
    if not user:
        return jsonify({
            'error': 'Certificate ID not found',
            'certificate_id': cert_id,
            'status': 'invalid'
        }), 404
    
    archive_path = f"archive/{cert_id}.pdf"
    
    if not os.path.exists(archive_path):
        return jsonify({
            'error': 'Certificate PDF not generated yet',
            'certificate_id': cert_id,
            'recipient_name': f"{user.first_name} {user.surname}",
            'status': user.status,
            'generated': False,
            'help': 'Certificate may still be processing or failed. Check worker logs.'
        }), 404
    
    return send_file(
        archive_path,
        mimetype='application/pdf',
        as_attachment=False,
        download_name=f'{cert_id}.pdf'
    )

@bp.route('/api/archive/status/<certificate_id>')
def archive_status(certificate_id):
    """
    Quick status check — does the PDF exist in archive?
    """
    from flask import jsonify
    user = User.query.filter_by(certificate_id=certificate_id).first()
    
    if not user:
        return jsonify({'error': 'Invalid certificate ID'}), 404
    
    archive_path = f"archive/{certificate_id}.pdf"
    pdf_exists = os.path.exists(archive_path)
    
    return jsonify({
        'certificate_id': certificate_id,
        'recipient': f"{user.first_name} {user.surname}",
        'email': user.email,
        'status': user.status,
        'pdf_generated': pdf_exists,
        'pdf_size_bytes': os.path.getsize(archive_path) if pdf_exists else 0,
        'verification_url': f"/api/archive/view/{certificate_id}"
    })


@bp.route('/api/archive/purge', methods=['POST'])
@login_required
def purge_archive():
    """Delete certificate PDFs from archive. Frees storage while keeping DB records."""
    data = request.json or {}
    cert_type_id = data.get('cert_type_id')
    purge_all = data.get('purge_all', False)

    if purge_all:
        cert_types = CertificateType.query.all()
    elif cert_type_id:
        ct = CertificateType.query.get(cert_type_id)
        if not ct:
            return jsonify({'error': 'Certificate type not found'}), 404
        cert_types = [ct]
    else:
        return jsonify({'error': 'Specify cert_type_id or purge_all=true'}), 400

    purged_count = 0
    freed_bytes = 0

    for ct in cert_types:
        users = User.query.filter_by(certificate_type_id=ct.id).all()
        for user in users:
            if user.certificate_id:
                archive_path = f"archive/{user.certificate_id}.pdf"
                if os.path.exists(archive_path):
                    freed_bytes += os.path.getsize(archive_path)
                    os.remove(archive_path)
                    purged_count += 1

    db.session.add(AuditLog(
        action='archive_purged',
        performed_by=current_user.email,
        details={'purged': purged_count, 'freed_bytes': freed_bytes}
    ))
    db.session.commit()

    return jsonify({
        'message': f'Purged {purged_count} certificate PDFs',
        'freed_mb': round(freed_bytes / (1024 * 1024), 2),
        'freed_bytes': freed_bytes
    })


@bp.route('/api/archive/stats')
@login_required
def archive_stats():
    """Return archive storage usage stats."""
    archive_dir = 'archive'
    if not os.path.exists(archive_dir):
        return jsonify({'total_files': 0, 'total_size_mb': 0, 'total_size_bytes': 0})

    total_size = 0
    file_count = 0
    for filename in os.listdir(archive_dir):
        if filename.endswith('.pdf'):
            fp = os.path.join(archive_dir, filename)
            total_size += os.path.getsize(fp)
            file_count += 1

    return jsonify({
        'total_files': file_count,
        'total_size_mb': round(total_size / (1024 * 1024), 2),
        'total_size_bytes': total_size
    })

